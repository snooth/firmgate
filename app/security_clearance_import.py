"""Parse security clearance spreadsheet rows (Excel upload or JSON)."""

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

_MONTHS = {
    "jan": "01",
    "feb": "02",
    "mar": "03",
    "apr": "04",
    "may": "05",
    "jun": "06",
    "jul": "07",
    "aug": "08",
    "sep": "09",
    "oct": "10",
    "nov": "11",
    "dec": "12",
}


def _norm_header(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _pick(row: dict[str, Any], keys: list[str]) -> str:
    lower: dict[str, Any] = {}
    for k, v in row.items():
        nk = _norm_header(k)
        if nk:
            lower[nk] = v
    for k in keys:
        v = lower.get(_norm_header(k))
        if v is None:
            continue
        if isinstance(v, (datetime, date)):
            return v
        s = str(v).strip()
        if s and s.lower() not in ("none", "null", "n/a"):
            return s
    return ""


def _excel_serial_to_iso(n: float) -> str:
    try:
        v = float(n)
    except (TypeError, ValueError):
        return ""
    if v <= 0:
        return ""
    ms = int(round((v - 25569) * 86400 * 1000))
    try:
        return datetime.utcfromtimestamp(ms / 1000).strftime("%Y-%m-%d")
    except (OSError, OverflowError, ValueError):
        return ""


def _to_iso_date(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        iso = _excel_serial_to_iso(float(v))
        if iso:
            return iso
    s = str(v).strip()
    if not s:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", s)
    if m:
        dd, mm, yy = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        return f"{yy}-{mm}-{dd}"
    m = re.match(r"^(\d{2})-([A-Za-z]{3})-(\d{4})$", s)
    if m:
        mm = _MONTHS.get(m.group(2).lower(), "")
        if mm:
            return f"{m.group(3)}-{mm}-{m.group(1)}"
    try:
        t = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return t.date().isoformat()
    except ValueError:
        pass
    try:
        t = datetime.strptime(s, "%d/%m/%Y")
        return t.date().isoformat()
    except ValueError:
        return ""


def _normalize_level(raw: str) -> str:
    s = str(raw or "").strip()
    if not s:
        return ""
    up = s.upper()
    if up == "BASELINE":
        return "Baseline"
    if up in ("NV1", "NV2", "PV"):
        return up
    return s


def _derive_status(expiry_iso: str, explicit: str) -> str:
    if explicit:
        return explicit
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", expiry_iso):
        return "Active"
    try:
        exp = date.fromisoformat(expiry_iso)
    except ValueError:
        return "Active"
    today = date.today()
    if exp < today:
        return "Expired"
    if (exp - today).days <= 90:
        return "Expiring soon"
    return "Active"


def row_from_mapping(raw: dict[str, Any]) -> dict[str, Any] | None:
    csid = _pick(
        raw,
        [
            "csid",
            "cs id",
            "cs_id",
            "cs id.",
            "csid.",
            "personnel id",
            "personnel number",
            "employee id",
            "employee number",
            "id",
            "clearance id",
        ],
    )
    if not csid:
        return None
    # Strip Excel float artifacts (762056.0 -> CS762056 if needed)
    if re.match(r"^\d+(\.0+)?$", csid):
        csid = str(int(float(csid)))
    def _date_field(keys: list[str], *extra_keys: str) -> str:
        picked = _pick(raw, keys)
        if picked != "":
            return _to_iso_date(picked)
        for ek in extra_keys:
            if ek in raw and raw[ek] is not None:
                return _to_iso_date(raw[ek])
        return ""

    revalidation = _date_field(
        ["revalidation date", "revalidation", "revalidation_date", "re-validation date"],
        "revalidation",
    )
    grant_date = _date_field(
        [
            "security clearance grant date",
            "clearance grant date",
            "grant date",
            "grant_date",
            "clearance granted",
        ],
        "grant_date",
    )
    expiry = _date_field(
        ["expiry date", "expiry", "expiry_date", "expiration date", "expiration", "expire date"],
        "expiry",
    )
    level = _normalize_level(
        _pick(raw, ["clearance level", "level", "clearance", "clearance_level"]) or raw.get("level")
    )
    explicit_status = _pick(raw, ["status"]) or str(raw.get("status") or "").strip()
    return {
        "csid": csid[:120],
        "given": _pick(raw, ["given name", "given", "first name", "firstname", "first"])[:200],
        "family": _pick(
            raw, ["family name", "family", "last name", "lastname", "surname", "last"]
        )[:200],
        "agent_request_from": _pick(
            raw,
            [
                "agent/request from",
                "agent request from",
                "agent",
                "request from",
                "agent_request_from",
            ],
        )[:120],
        "level": level[:40],
        "revalidation": revalidation[:32],
        "grant_date": grant_date[:32],
        "expiry": expiry[:32],
        "status": _derive_status(expiry, explicit_status)[:40],
        "archived": False,
    }


def parse_workbook_bytes(data: bytes, filename: str = "") -> list[dict[str, Any]]:
    """Read first sheet of .xlsx/.xlsm into row dicts for import."""
    name = str(filename or "").lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        raise ValueError("Legacy .xls files: save as .xlsx in Excel, then import again.")
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError("openpyxl is not installed on the server.") from e

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return []
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = [_norm_header(c) for c in header_row]
        out: list[dict[str, Any]] = []
        for cells in rows_iter:
            if cells is None:
                continue
            raw: dict[str, Any] = {}
            for i, h in enumerate(headers):
                if not h or i >= len(cells):
                    continue
                raw[h] = cells[i]
            if not any(v is not None and str(v).strip() for v in raw.values()):
                continue
            parsed = row_from_mapping(raw)
            if parsed:
                out.append(parsed)
            if len(out) >= 5000:
                break
        return out
    finally:
        wb.close()


def parse_json_record_list(raw: list) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    if not isinstance(raw, list):
        return out
    for item in raw[:5000]:
        if not isinstance(item, dict):
            continue
        if "csid" in item and item.get("csid"):
            out.append(
                {
                    "csid": str(item.get("csid") or "").strip(),
                    "given": str(item.get("given") or "").strip(),
                    "family": str(item.get("family") or "").strip(),
                    "agent_request_from": str(item.get("agent_request_from") or "").strip(),
                    "level": str(item.get("level") or "").strip(),
                    "revalidation": str(item.get("revalidation") or "").strip(),
                    "grant_date": str(item.get("grant_date") or "").strip(),
                    "expiry": str(item.get("expiry") or "").strip(),
                    "status": str(item.get("status") or "").strip(),
                    "archived": bool(item.get("archived")),
                }
            )
            continue
        parsed = row_from_mapping(item)
        if parsed:
            out.append(parsed)
    return out
